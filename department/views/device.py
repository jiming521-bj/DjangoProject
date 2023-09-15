# -*- coding: utf-8 -*-
# @Time     : 2023/9/12 19:03
# @Author   : JiMing
# @File     : device.py
# @SoftWare : PyCharm

from django.shortcuts import render, HttpResponse, redirect
from django.views.decorators.csrf import csrf_exempt
from department.my_forms import Device_List, DeviceEditModelForm
from django.http import JsonResponse  # 返回JSON格式的数据
import json
from department import models
from department.utils import myPage
from openpyxl import load_workbook


def device_list(request):
    """
    设备管理
    :param request:
    :return:
    """
    # 获取数据库中所有的任务信息
    querySet = models.Device.objects.all()
    # print(request.GET)
    # 分页显示
    page_object = myPage.pagination(request, querySet, page_size=4)
    # 创建Form对象
    forms = Device_List()
    context = {
        'forms': forms,
        'querySet': page_object.pageQuerySet,
        'page_html': page_object.html()
    }
    return render(request, 'device-list.html', context)


@csrf_exempt
def device_ajax(request):
    request_param = request.GET
    print(request_param)

    request_method_post = request.POST
    print(request_method_post)

    # print(request_param.get('name'))
    # print(request_param.get('age'))

    data_list = {
        'name': 'ming',
        'age': 22,
        'gender': '男',
        'hobby': 'basketball',
        'remark': '这是一场没有尽头的旅途'
    }

    # if request_param.get('name') != 'ming':
    #     return HttpResponse('失败了')
    return HttpResponse(json.dumps(data_list))


@csrf_exempt
def device_add(request):
    """
    添加任务
    :param request:
    :return:
    """
    forms = Device_List(data=request.POST)
    if forms.is_valid():
        forms.save()
        data_list = {
            'static': True
        }
        return HttpResponse(json.dumps(data_list))
    # print(forms.errors)
    data_list = {
        'static': False,
        'error': forms.errors
    }
    return HttpResponse(json.dumps(data_list))


def device_delete(request, nid):
    """
    删除
    :param nid:
    :param request:
    :return:
    """
    if request.method == 'GET':
        if not models.Device.objects.filter(id=nid).exists():
            return HttpResponse("<script>alert('数据不存在');window.location='/device/list/'</script>")
        models.Device.objects.filter(id=nid).delete()
        return redirect('/device/list/')


def device_muilt(request):
    """
    文件上传操作
    :param request:
    :return:
    """

    # 获取文件上传的内容
    files = request.FILES.get('files')
    files_name_list = ['xls', 'xlsx', 'xlsm', 'pdf', 'et']
    if not files:
        return HttpResponse("<script>alert('请上传文件');window.location='/device/list/';</script>")
    # 获取文件名
    f_name = files.name
    # 使用split获取文件后缀
    suffix_file = f_name.split('.')[1]

    # 排出不是Excel格式的文件
    if suffix_file not in files_name_list:
        return HttpResponse("<script>alert('请上传Excel格式的文件');window.location='/device/list/';</script>")

    # 处理Excel数据 将数据写入到Device数据库中
    # 1、获取Excel操作对象
    wb = load_workbook(files)

    # 2、获取工作簿 首个工作簿
    sheet = wb.worksheets[0]

    flag = True
    # 解析工作簿中的数据
    for i in range(2, sheet.max_row):
        # for j in range(0, sheet.max_column):
        # print(sheet[i][j].value)
        # 获取数据操作对象
        print(sheet[i][0])
        if models.Device.objects.filter(title=sheet[i][0].value).exists():
            flag = False
            continue
        try:
            models.Device.objects.create(
                title=sheet[i][0].value,
                detail=sheet[i][1].value,
                level=sheet[i][2].value,
                person_id=request.session['info'].get('id')
            )
        except ValueError:
            raise ValueError
    if not flag:
        return HttpResponse("<script>alert('列表中已经存在任务了，请不要重复添加！');window.location='/device/list/';</script>")
    # 返回响应数据
    return HttpResponse("<script>alert('文件上传成功');window.location='/device/list/';</script>")


def device_detail(request):
    """
    获取数据
    :param request:
    :return:
    """
    nid = request.GET.get('nid')
    if not models.Device.objects.filter(id=nid).exists():
        context = {
            'static': False,
            'error': '该条数据不存在',
        }
    values = models.Device.objects.filter(id=nid).values('title', 'detail', 'level', 'person').first()
    print(values)
    context = {
        'static': True,
        'values': values,
    }
    return HttpResponse(json.dumps(context))


@csrf_exempt
def device_edit(request):
    """
    修改任务管理
    :param request:
    :return:
    """
    nid = request.GET.get('nid')
    print(nid)
    exist = models.Device.objects.filter(id=nid).first()
    print(exist)

    if not exist:
        context = {
            'static': False,
            'error': '修改失败，数据不存在'
        }
        return JsonResponse(context)
    forms = DeviceEditModelForm(data=request.POST, instance=exist)
    if forms.is_valid():
        print(forms.cleaned_data)
        forms.save()
        return JsonResponse({'static': True})
    context = {
        'static': False,
        'error': forms.errors
    }
    return JsonResponse(context)
